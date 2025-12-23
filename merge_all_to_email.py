import pandas as pd
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
import psycopg2
from zoneinfo import ZoneInfo
from typing import Optional, Iterable, Tuple, List, Dict
from psycopg2 import Error
from psycopg2 import sql
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import time
from dotenv import load_dotenv

SEOUL_TZ = ZoneInfo("Asia/Seoul")

USE_DB = os.getenv("CRAWLING_USE_DB", "true").lower() in ("1", "true", "yes")  
SEND_EMAIL = os.getenv("CRAWLING_SEND_EMAIL", "true").lower() in ("1", "true", "yes")
REPORT_LIMIT = 50

COL_RANK = "Rank"
COL_BRAND = "Brand"
COL_PRODUCT_NAME = "Product"
COL_PRICE = "Price"
COL_DATETIME_TEXT = "Datetime"
COL_CHANNEL = "Channel"
COL_OLD_PRICE = "Old_Price"
COL_BRAND_KEY = "_brand_key"
COL_PRODUCT_KEY = "_product_key"

COL_COLLECTED_AT = "수집일시"
COL_PREVIOUS_RANK = "전일 순위"
COL_RANK_DELTA = "전일 변동"
COL_STATUS = "전일대비 증감"
COL_PREVIOUS_RANK_WEEK = "전주 순위"
COL_WEEK_DELTA = "전주 변동"
COL_WEEK_STATUS = "전주대비 증감"

load_dotenv(override=True)

CHANNEl_CONFIGS = [{
    "Channel": "jolse",
    "db":{
        "host":os.getenv("DB_HOST"),
        "port":os.getenv("DB_PORT"),
        "dbname":os.getenv("DB_DATABASE"),
        "user":os.getenv("DB_USER"),
        "password":os.getenv("DB_PASSWORD")
    },
},
{
    "Channel":"yesstyle",
    "db":{
        "host":os.getenv("DB_HOST"),
        "port":os.getenv("DB_PORT"),
        "dbname":os.getenv("DB_DATABASE"),
        "user":os.getenv("DB_USER"),
        "password":os.getenv("DB_PASSWORD")
    },
},
{
    "Channel":"stylevana",
    "db":{
        "host":os.getenv("DB_HOST"),
        "port":os.getenv("DB_PORT"),
        "dbname":os.getenv("DB_DATABASE"),
        "user":os.getenv("DB_USER"),
        "password":os.getenv("DB_PASSWORD")
    },
}]


VOLUME_PREFIX_PATTERN = re.compile(r"^\(\d+(?:ML|EA|G|PATCHES|PCS|PADS)\)\s*", re.IGNORECASE)


def _normalize_product_name(name: str) -> str:
    """용량 접두사/중복 공백을 제거해 비교 가능한 상품명을 만든다."""
    if name is None:
        return ""
    text = str(name).strip()
    text = VOLUME_PREFIX_PATTERN.sub("", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_match_key(value: Optional[str]) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", "", text)


def _add_match_keys_inplace(df: pd.DataFrame) -> pd.DataFrame:
    """브랜드/제품명 컬럼으로 비교용 키를 생성한다."""
    if df is None:
        return df
    if COL_BRAND in df.columns:
        df[COL_BRAND_KEY] = df[COL_BRAND].apply(_normalize_match_key)
    else:
        df[COL_BRAND_KEY] = ""
    if COL_PRODUCT_NAME in df.columns:
        df[COL_PRODUCT_KEY] = df[COL_PRODUCT_NAME].apply(_normalize_match_key)
    else:
        df[COL_PRODUCT_KEY] = ""
    return df


def _preprocess_rank_dataframe(df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    """순위/브랜드/상품명을 기본 정제해 비교 가능한 DataFrame으로 만든다."""
    if df is None:
        return None
    work = df.copy()
    if COL_PRODUCT_NAME in work.columns:
        work = work[work[COL_PRODUCT_NAME].notna()]
        work[COL_PRODUCT_NAME] = work[COL_PRODUCT_NAME].apply(_normalize_product_name)
    if COL_RANK in work.columns:
        work[COL_RANK] = pd.to_numeric(work[COL_RANK], errors="coerce")
        work = work[work[COL_RANK].notna()]
        work[COL_RANK] = work[COL_RANK].astype(int)
    for col in (COL_BRAND, COL_CHANNEL):
        if col in work.columns:
            work[col] = (
                work[col]
                .fillna("")
                .apply(lambda x: re.sub(r"\s+", " ", str(x)).strip())
            )
    _add_match_keys_inplace(work)
    return work


def fetch_channel_data(db_cfg,Channel, start_ts, end_ts):
    conn = psycopg2.connect(**db_cfg)
    with conn, conn.cursor() as cursor:
        schema_identifier = sql.Identifier("suncream_crawling")
        table_identifier = sql.Identifier(f"{Channel}")
        query = sql.SQL(
            """
            SELECT "Rank", "Brand", "Product", "Old_price", "Price", "DateTime", "Channel"
            FROM {}.{}
            WHERE "Channel" = %s                                                                                                                                                                                           
            AND "DateTime" >= %s                                                                                                                                                                                         
            AND "DateTime" < %s                                                                                                                                                                                          
            ORDER BY "Rank"                                                                                                                                                                                                
            """
        ).format(schema_identifier, table_identifier)
        cursor.execute(query,(Channel,start_ts,end_ts))
        rows = cursor.fetchall()
        columns = ["Rank", "Brand", "Product", "Old_price", "Price", "DateTime", "Channel"]
    return pd.DataFrame(rows, columns=columns)


def fetch_previous_snapshot_from_db(
    Channel: str,
    target_time: datetime,
    limit: Optional[int] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[datetime], Optional[str]]:
    if not USE_DB:
        print("DB 조회 비활성화(stylevana_USE_DB=false); 이전 스냅샷 없이 진행합니다.")
        return None, None, None

    connection = None
    try:
        connection = psycopg2.connect(
            host=os.getenv("DB_HOST"),
            port=os.getenv("DB_PORT"),
            dbname=os.getenv("DB_DATABASE"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
        )
        cursor = connection.cursor()

        target_time = target_time.replace(tzinfo=None)
        desired_time = target_time.replace(hour=8, minute=0, second=0, microsecond=0)
        start_ts = desired_time.replace(hour=0, minute=0, second=0, microsecond=0)
        end_ts = start_ts + timedelta(days=1)

        schema_identifier = sql.Identifier(Channel)
        table_identifier = sql.Identifier(f"{Channel}_table")

        snapshot_query = sql.SQL(
            """
            SELECT created_at
            FROM {}.{}
            WHERE "Channel" = %s
            AND "DateTime" >= %s
            AND "DateTime" < %s
            ORDER BY created_at DESC
            """
        ).format(schema_identifier, table_identifier)
        cursor.execute(snapshot_query, (Channel, start_ts, end_ts))

        day_rows = [row[0] for row in cursor.fetchall() if row[0] is not None]
        if day_rows:
            latest_created = min(day_rows, key=lambda ts: abs(ts - desired_time))
        else:
            fallback_query = sql.SQL(
                """
                SELECT MAX(created_at)
                FROM {}.{}
                WHERE "Channel" = %s
                AND created_at < %s
                """
            ).format(schema_identifier, table_identifier)
            cursor.execute(fallback_query, (Channel, target_time))
            result = cursor.fetchone()
            latest_created = result[0] if result else None
            if latest_created is None:
                return None, None, None

        base_query = sql.SQL(
            """
            SELECT
                "Rank",
                "Brand",
                "Product",
                "Old_price",
                "Price",
                "DateTime",
                "Channel",
                created_at
            FROM {}.{}
            WHERE "Channel" = %s
            AND created_at = %s
            ORDER BY "Rank" ASC
            """
        ).format(schema_identifier, table_identifier)

        params: List = [Channel, latest_created]
        if limit is not None:
            base_query = base_query + sql.SQL(" LIMIT %s")
            params.append(limit)

        cursor.execute(base_query, params)
        rows = cursor.fetchall()
        if not rows:
            return None, latest_created, None

        df = pd.DataFrame(
            rows,
            columns=[
                COL_RANK,
                COL_BRAND,
                COL_PRODUCT_NAME,
                COL_OLD_PRICE,
                COL_PRICE,
                COL_DATETIME_TEXT,
                COL_CHANNEL,
                COL_COLLECTED_AT,
            ],
        )
        df[COL_RANK] = pd.to_numeric(df[COL_RANK], errors="coerce")
        df = df.dropna(subset=[COL_RANK])
        df[COL_RANK] = df[COL_RANK].astype(int)
        df.sort_values(by=COL_RANK, inplace=True)
        df.reset_index(drop=True, inplace=True)

        label = latest_created.strftime("%Y-%m-%d %H:%M") if latest_created else None

        return df, latest_created, label
    except Error as exc:
        print(f"PostgreSQL 조회 오류: {exc}")
        return None, None, None
    finally:
        if connection:
            connection.close()


# 전날, 전주 비교
def _format_status_from_delta(delta_value) -> str:
    if pd.isna(delta_value):
        return "신규"
    if delta_value > 0:
        return f"▲{int(delta_value)}"
    if delta_value < 0:
        return f"▼{abs(int(delta_value))}"
    return "-"


def annotate_rank_changes(
    current_df: pd.DataFrame,
    previous_df: Optional[pd.DataFrame],
) -> Tuple[pd.DataFrame, Optional[pd.DataFrame], Optional[str]]:
    """전일 데이터와 비교해 순위 변동/차트아웃을 계산합니다."""
    current_clean = _preprocess_rank_dataframe(current_df)
    previous_clean = _preprocess_rank_dataframe(previous_df)

    if previous_clean is None or previous_clean.empty:
        annotated = current_clean.copy()
        annotated[COL_PREVIOUS_RANK] = pd.NA
        annotated[COL_RANK_DELTA] = pd.NA
        annotated[COL_STATUS] = "신규"
        annotated.sort_values(by=[COL_RANK], inplace=True)
        return annotated, None, None

    keys = [COL_CHANNEL, COL_BRAND_KEY, COL_PRODUCT_KEY]
    merged = current_clean.merge(
        previous_clean[keys + [COL_RANK]],
        how="left",
        on=keys,
        suffixes=("", "_previous"),
    )

    merged.rename(columns={f"{COL_RANK}_previous": COL_PREVIOUS_RANK}, inplace=True)
    merged[COL_PREVIOUS_RANK] = pd.to_numeric(
        merged[COL_PREVIOUS_RANK], errors="coerce"
    ).astype("Int64")
    merged[COL_RANK_DELTA] = merged[COL_PREVIOUS_RANK] - merged[COL_RANK]
    merged[COL_RANK_DELTA] = pd.to_numeric(
        merged[COL_RANK_DELTA], errors="coerce"
    ).astype("Int64")

    merged[COL_STATUS] = merged[COL_RANK_DELTA].apply(_format_status_from_delta)

    dropped = previous_clean.merge(
        merged[keys],
        how="left",
        on=keys,
        indicator=True,
    )
    dropped = dropped[dropped["_merge"] == "left_only"].drop(columns=["_merge"])
    if dropped.empty:
        dropped_df = None
    else:
        dropped_df = dropped.sort_values(by=[COL_RANK]).copy()
        dropped_df[COL_PREVIOUS_RANK] = dropped_df[COL_RANK]
        dropped_df.drop(columns=[COL_RANK], inplace=True)
        dropped_df[COL_STATUS] = "차트 아웃"

    previous_label = None
    if previous_df is not None and COL_DATETIME_TEXT in previous_df.columns:
        series = previous_df[COL_DATETIME_TEXT].dropna()
        if not series.empty:
            previous_label = series.iloc[0]

    merged.sort_values(by=[COL_RANK], inplace=True)
    return merged, dropped_df, previous_label

def add_weekly_rank_changes(
    current_df: pd.DataFrame,
    previous_week_df: Optional[pd.DataFrame],
) -> Tuple[pd.DataFrame, Optional[str]]:
    """전주 데이터와 비교해 주간 변동을 계산합니다."""
    if previous_week_df is None or previous_week_df.empty:
        df = current_df.copy()
        df[COL_PREVIOUS_RANK_WEEK] = pd.NA
        df[COL_WEEK_DELTA] = pd.NA
        df[COL_WEEK_STATUS] = "-"
        return df, None

    prev_clean = _preprocess_rank_dataframe(previous_week_df)
    if prev_clean is None or prev_clean.empty:
        df = current_df.copy()
        df[COL_PREVIOUS_RANK_WEEK] = pd.NA
        df[COL_WEEK_DELTA] = pd.NA
        df[COL_WEEK_STATUS] = "-"
        return df, None

    keys = [COL_CHANNEL, COL_BRAND_KEY, COL_PRODUCT_KEY]
    merged = current_df.merge(
        prev_clean[keys + [COL_RANK]],
        how="left",
        on=keys,
        suffixes=("", "_week"),
    )

    merged.rename(columns={f"{COL_RANK}_week": COL_PREVIOUS_RANK_WEEK}, inplace=True)
    merged[COL_PREVIOUS_RANK_WEEK] = pd.to_numeric(merged[COL_PREVIOUS_RANK_WEEK], errors="coerce").astype("Int64")
    merged[COL_WEEK_DELTA] = merged[COL_PREVIOUS_RANK_WEEK] - merged[COL_RANK]
    merged[COL_WEEK_DELTA] = pd.to_numeric(merged[COL_WEEK_DELTA], errors="coerce").astype("Int64")
    merged[COL_WEEK_STATUS] = merged[COL_WEEK_DELTA].apply(_format_status_from_delta)

    week_label = None
    if previous_week_df is not None and COL_DATETIME_TEXT in previous_week_df.columns:
        series = previous_week_df[COL_DATETIME_TEXT].dropna()
        if not series.empty:
            week_label = series.iloc[0]

    merged.sort_values(by=[COL_RANK], inplace=True)
    return merged, week_label
def start_email(run_time: datetime) -> str:
    lines: List[any] = [
        "안녕하세요 위마케팅 사업지원팀 김희교입니다.",
        f"금일자 채널별 선크림 전일·전주 대비 랭킹 변동 현황 보고드립니다.",
        "",
        f"수집 일시 (KST): {run_time.strftime('%Y-%m-%d %H:%M')}",
        ""
    ]
    return "\n".join(lines)

def end_email():
    lines : List[str] = [
        "감사합니다.",
        "김희교 드림"
    ]

    return "\n".join(lines)

def summarize_changes_for_email(
    Channel : str,
    annotated_df: pd.DataFrame,
    dropped_df: Optional[pd.DataFrame],
    run_time: datetime,
    previous_label: Optional[str],
    previous_week_label: Optional[str],
) -> str:
    # lines: List[str] = [
    #     "안녕하세요 위마케팅 사업지원팀 김희교입니다.",
    #     f"금일자 {Channel}별 선크림 전일·전주 대비 랭킹 변동 현황 보고드립니다.",
    #     "",
    #     f"수집 일시 (KST): {run_time.strftime('%Y-%m-%d %H:%M')}",
    # ]
    lines : List[str] = []
    if previous_label:
        lines.append(f"전일 기준: {previous_label}")
    else:
        lines.append("전일 기준: 참고 가능한 데이터 없음")
    if previous_week_label:
        lines.append(f"전주 기준: {previous_week_label}")
    else:
        lines.append("전주 기준: 참고 가능한 데이터 없음")
    lines.append("")

    for category in annotated_df[COL_CHANNEL].dropna().unique():
        cat_df = annotated_df[annotated_df[COL_CHANNEL] == category].copy()
        cat_df = cat_df[cat_df[COL_RANK] <= REPORT_LIMIT]
        cat_df["_delta"] = pd.to_numeric(cat_df[COL_RANK_DELTA], errors="coerce")

        up_count = int((cat_df["_delta"] > 0).sum())
        down_count = int((cat_df["_delta"] < 0).sum())
        new_count = int((cat_df[COL_STATUS] == "신규").sum())
        keep_count = int((cat_df[COL_STATUS] == "-").sum())

        lines.append(f"[{category}] ▲ {up_count} / ▼ {down_count} / 신규 {new_count} / 유지 {keep_count}")

        highlights: List[str] = []
        seen: set[Tuple[str, str]] = set()

        def _append_rows(df: pd.DataFrame):
            for _, r in df.iterrows():
                brand = r.get(COL_BRAND, "")
                product = r.get(COL_PRODUCT_NAME, "")
                key = (brand, product)
                status = str(r.get(COL_STATUS, "") or "")
                if key in seen or not status or status == "-":
                    continue
                highlights.append(f"  {status}: {brand} - {product}(#{int(r[COL_RANK])})")
                seen.add(key)
                if len(highlights) >= 7:
                    break

        cat_df["_abs_delta"] = cat_df["_delta"].abs()
        primary = cat_df[cat_df["_abs_delta"] >= 2].sort_values(by=["_abs_delta", COL_RANK], ascending=[False, True])
        _append_rows(primary)

        if len(highlights) < 5:
            brand_mask = cat_df[COL_BRAND].fillna("").str.lower()
            priority = cat_df[
                brand_mask.isin({"numbuzin", "fwee"})
                & (
                    cat_df["_abs_delta"].fillna(0) > 0
                    | (cat_df[COL_STATUS] == "신규")
                )
            ].sort_values(by=["_abs_delta", COL_RANK], ascending=[False, True])
            _append_rows(priority)

        if len(highlights) < 7:
            top10 = cat_df[(cat_df[COL_RANK] <= 10) & (cat_df[COL_STATUS] != "-")].sort_values(by=COL_RANK)
            _append_rows(top10)

        lines.extend(highlights)
        if highlights:
            lines.append("")

    if dropped_df is not None and not dropped_df.empty:
        filtered = dropped_df[
            dropped_df[COL_PREVIOUS_RANK].notna()
            & (dropped_df[COL_PREVIOUS_RANK] <= REPORT_LIMIT)
        ].copy()
        if not filtered.empty:
            lines.append(f"차트 아웃 {len(filtered)}개")
            top_drop = filtered.sort_values(COL_PREVIOUS_RANK).head(5)
            for _, row in top_drop.iterrows():
                previous_rank = row.get(COL_PREVIOUS_RANK)
                rank_text = f"#{int(previous_rank)}" if pd.notna(previous_rank) else "-"
                lines.append(f"  {rank_text} {row.get(COL_BRAND, '')} - {row.get(COL_PRODUCT_NAME, '')}")
    return "\n".join(lines)

# 엑셀 시트 작성
def write_channel_sheet(
    writer,
    channel: str,
    annotated_df: pd.DataFrame,
    run_time: datetime,
    top_n: int = REPORT_LIMIT,
) -> None:
    report_df = (
        annotated_df.copy()
        .dropna(subset=[COL_RANK])
        .sort_values(by=[COL_RANK])
    )
    report_df = report_df.head(top_n)

    display_df = report_df[
        [
            COL_RANK,
            COL_BRAND,
            COL_PRODUCT_NAME,
            COL_PRICE,
            COL_STATUS,
            COL_WEEK_STATUS,
        ]
    ].copy()
    display_df.rename(
        columns={
            COL_RANK: "랭킹",
            COL_BRAND: "브랜드",
            COL_PRODUCT_NAME: "제품명",
            COL_PRICE: "할인가",
            COL_STATUS: "전일대비 증감",
            COL_WEEK_STATUS: "전주대비 증감",
        },
        inplace=True,
    )
    display_df["전일대비 증감"] = display_df["전일대비 증감"].fillna("-")
    display_df["전주대비 증감"] = display_df["전주대비 증감"].fillna("-")
    display_df["비고"] = ""

    sheet_name = channel
    start_row = 5
    start_col = 3

    display_df.to_excel(
        writer,
        index=False,
        sheet_name=sheet_name,
        startrow=start_row,
        startcol=start_col,
    )

    worksheet = writer.sheets[sheet_name]
    worksheet["D2"] = f"■ {channel} 랭킹 변동사항"
    worksheet["D2"].font = Font(bold=True)
    worksheet["D2"].alignment = Alignment(horizontal="left")

    worksheet["D4"] = f"- 기준일자 : {run_time.date()}"
    worksheet["D4"].alignment = Alignment(horizontal="left")
    header_row = start_row + 1
    header_font = Font(bold=True)
    for cells in worksheet.iter_rows(
        min_row=header_row,
        max_row=header_row,
        min_col=start_col + 1,
        max_col=start_col + len(display_df.columns),
    ):
        for cell in cells:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    data_row_start = header_row + 1
    for row in worksheet.iter_rows(
        min_row=data_row_start,
        max_row=data_row_start + len(display_df) - 1,
        min_col=start_col + 1,
        max_col=start_col + len(display_df.columns),
    ):
        for idx, cell in enumerate(row):
            if idx in (0, 4, 5):
                cell.alignment = Alignment(horizontal="center")
            elif idx == 3:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for offset in (4, 5):
        trend_col = get_column_letter(start_col + 1 + offset)
        for row_idx in range(data_row_start, data_row_start + len(display_df)):
            cell = worksheet[f"{trend_col}{row_idx}"]
            value = str(cell.value) if cell.value is not None else ""
            if value.startswith("▲"):
                cell.font = Font(color="FF0000")
            elif value.startswith("▼"):
                cell.font = Font(color="0000FF")

    column_widths = [8, 22, 40, 18, 12, 12, 18]
    for offset, width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(start_col + 1 + offset)].width = width

    worksheet.freeze_panes = f"{get_column_letter(start_col + 1)}{header_row + 1}"

def build_channel_report(
    channel_cfg: Dict[str, Dict[str, Optional[str]]],
    run_time: datetime,
) -> Optional[Dict[str, object]]:
    """단일 채널 데이터를 조회/비교해 보고서 작성에 필요한 데이터를 반환한다."""
    channel = channel_cfg["Channel"]
    db_cfg = channel_cfg["db"]

    seoul_now = run_time.astimezone(SEOUL_TZ)
    start_ts = seoul_now.replace(hour=0, minute=0, second=0, microsecond=0).replace(tzinfo=None)
    end_ts = start_ts + timedelta(days=1)

    try:
        current_df = fetch_channel_data(db_cfg, channel, start_ts, end_ts)
    except Exception as exc:
        print(f"[{channel}] 현재 데이터 조회 실패: {exc}")
        return None

    if current_df.empty:
        print(f"[{channel}] 조회된 데이터가 없어 보고서를 건너뜁니다.")
        return None

    prev_day_time = run_time - timedelta(days=1)
    prev_week_time = run_time - timedelta(days=7)

    previous_df, _, _ = fetch_previous_snapshot_from_db(channel, prev_day_time)
    annotated_df, dropped_df, previous_label = annotate_rank_changes(current_df, previous_df)

    previous_week_df, _, _ = fetch_previous_snapshot_from_db(channel, prev_week_time)
    annotated_with_week, week_label = add_weekly_rank_changes(annotated_df, previous_week_df)

    if previous_label:
        print(f"[{channel}] 전일 비교 기준: {previous_label}")
    if week_label:
        print(f"[{channel}] 전주 비교 기준: {week_label}")
    if dropped_df is not None and not dropped_df.empty:
        print(f"[{channel}] 차트 아웃 {len(dropped_df)}건")

    return {
        "channel": channel,
        "annotated_df": annotated_with_week,
        "dropped_df": dropped_df,
        "previous_label": previous_label,
        "week_label": week_label,
    }

# 메일 보내기
def send_email_from_memory(Channel: str, excel_bytes: bytes, file_name: str, body_text: Optional[str] = None):
    """
    CC/BCC는 환경변수로 제어한다.
    - 우선순위: `<CHANNEL>_CC` → `RANKING_CC` 
    - 구분자: , 또는 ;  (공백 허용)
    - 토큰 접두사:
        * 'cc:'  → 참조
        * 'bcc:' → 숨은참조
      접두사가 없으면 기본적으로 CC로 처리.
    예)
      RANKING_CC="cc:lead@domain.com, bcc:boss@domain.com;ops@domain.com"
      → CC=[lead@, ops@], BCC=[boss@]
    """
    def _split(addr_string: Optional[str]) -> List[str]:
        if not addr_string:
            return []
        return [a.strip() for a in re.split(r"[;,]", addr_string) if a.strip()]

    def _parse_cc_tokens(value: Optional[str]) -> Tuple[List[str], List[str]]:
        cc_list: List[str] = []
        bcc_list: List[str] = []
        for token in _split(value):
            lower = token.lower()
            if lower.startswith("bcc:"):
                addr = token[4:].strip()
                if addr:
                    bcc_list.append(addr)
            elif lower.startswith("cc:"):
                addr = token[3:].strip()
                if addr:
                    cc_list.append(addr)
            else:
                # 접두사 없으면 CC로 간주
                cc_list.append(token)
        # 중복 제거(순서 유지)
        cc_list = list(dict.fromkeys(cc_list))
        bcc_list = list(dict.fromkeys(bcc_list))
        return cc_list, bcc_list

    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    sender_email = os.getenv("EMAIL_ADDRESS") or os.getenv("SENDER_EMAIL")
    sender_password = os.getenv("EMAIL_PASSWORD") or os.getenv("SENDER_PASSWORD")

    # 받는 사람(TO) — 기존 변수를 유지
    to_list = _split(
        os.getenv("EMAIL_RECIPIENTS")
        or os.getenv("EMAIL_RECIPIENT")
        or os.getenv("RECIPIENT_EMAIL")
        or "diana0305@wemarketing.co.kr"
    )

    channel_env_key = f"{Channel.upper()}_CC"
    cc_env = (
        os.getenv(channel_env_key)
        or os.getenv("RANKING_CC")
    )
    cc_list, bcc_list = _parse_cc_tokens(cc_env)

    if not sender_email or not sender_password or not (to_list or cc_list or bcc_list):
        print("Email credentials/recipients are missing; skipping email send.")
        return

    msg = MIMEMultipart()
    msg["From"] = sender_email
    if to_list:
        msg["To"] = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    subject_date = datetime.now(SEOUL_TZ).strftime("%Y%m%d")  # KST 기준 오늘 날짜
    msg["Subject"] = f"{Channel} 랭킹 리포트_{subject_date}"

    body = body_text or f"{Channel} crawling completed.\nPlease see the attached report: {file_name}"
    msg.attach(MIMEText(body, "plain", _charset="utf-8"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={file_name}")
    msg.attach(part)

    # 실제 전송 대상: To + Cc + Bcc (중복 제거)
    all_recipients = list(dict.fromkeys([*to_list, *cc_list, *bcc_list]))

    attempts = 2
    for attempt in range(1, attempts + 1):
        try:
            server = smtplib.SMTP(smtp_server, smtp_port, timeout=30)
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, all_recipients, msg.as_string())
            server.quit()
            print(f"Email sent successfully to {', '.join(all_recipients)}.")
            return
        except Exception as exc:
            print(f"Email send failed (attempt {attempt}/{attempts}): {exc}")
            if attempt < attempts:
                print("Retrying in 60 seconds...")
                time.sleep(60)
    print("Giving up on email send after retries.")

def main():
    run_time = datetime.now(SEOUL_TZ)
    output_dir = Path(__file__).resolve().parent / "data"
    output_dir.mkdir(parents=True, exist_ok=True)

    channel_results: List[Dict[str, object]] = []

    for channel_cfg in CHANNEl_CONFIGS:
        channel = channel_cfg["Channel"]
        try:
            result = build_channel_report(channel_cfg, run_time)
            if result:
                channel_results.append(result)
        except Exception as exc:
            print(f"[{channel}] 리포트 생성 실패: {exc}")

    if not channel_results:
        print("생성할 리포트가 없습니다.")
        return

    timestamp_str = run_time.strftime("%Y%m%d_%H%M")
    report_path = output_dir / f"{timestamp_str}_ranking.xlsx"
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        for result in channel_results:
            channel = result["channel"]
            annotated_df = result["annotated_df"]
            write_channel_sheet(writer, channel, annotated_df, run_time)

    print(f"통합 엑셀 리포트 저장 완료: {report_path}")

    if SEND_EMAIL:
        body_sections = []
        body_sections.append(start_email(run_time))
        body_sections.append(end_email())
        email_body = "\n\n".join(body_sections)
        
        with report_path.open("rb") as fp:
            send_email_from_memory("ALL_CHANNELS", fp.read(), report_path.name, email_body)
    else:
        print("이메일 전송 비활성화(STYLEVANA_SEND_EMAIL=false)")


if __name__ == "__main__":
    main()


